import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

st.set_page_config(page_title="Sam Roasters - Auto Report", page_icon="☕", layout="centered")

st.title("☕ ระบบสร้างรายงานอัตโนมัติ")
st.markdown("โยนไฟล์รายงานของเมื่อวาน และไฟล์ POS ของวันนี้ ระบบจะหยอดข้อมูลลงช่องและคำนวณให้ทันที")

with st.form("upload_form"):
    st.info("📝 1. ไฟล์ตั้งต้น (Template)")
    template_file = st.file_uploader("โยนไฟล์รายงาน Excel ของเมื่อวาน (ไฟล์ที่มีฟอร์ม)", type=['xlsx'])
    
    st.info("📊 2. ไฟล์ข้อมูลจาก POS ของวันนี้")
    sale_file = st.file_uploader("ไฟล์ยอดรวมและส่วนลด (SaleReport) - ดึงเฉพาะ Discount", type=['csv', 'xlsx'])
    bev_file = st.file_uploader("ไฟล์จำนวนแก้ว (SaleByBehavior) - ดึงแก้วและยอด EatIn, Grab ฯลฯ", type=['csv', 'xlsx'])
    
    submit_button = st.form_submit_button("ประมวลผลและสร้างไฟล์ Excel 🚀")

def find_col_index(row_strs, possible_names):
    for i, val in enumerate(row_strs):
        if val.lower() in possible_names:
            return i
    return -1

if submit_button:
    if not template_file or not sale_file or not bev_file:
        st.error("❌ กรุณาโยนไฟล์ให้ครบทั้ง 3 ช่องครับ")
    else:
        with st.spinner("กำลังประมวลผล..."):
            try:
                data_map = {}
                log_messages = []

                # --- 1. สกัดข้อมูลจากไฟล์ SaleByBehavior ---
                df_bev = pd.read_csv(bev_file, header=None) if bev_file.name.endswith('.csv') else pd.read_excel(bev_file, header=None)
                
                item_names = ['สินค้า', 'menu', 'item', 'product']
                qty_names = ['จำนวน', 'qty', 'quantity']
                type_names = ['ปรเะภท', 'ประเภท', 'type', 'category', 'behavior']
                del_names = ['delivery', 'platform', 'app']
                amt_names = ['ยอดเงิน', 'amount', 'gross', 'gross sale', 'ยอดขาย'] # ใช้ยอดเงินเป็นหลัก
                net_names = ['ยอดสุทธิ', 'net sale', 'net', 'total']

                item_col = qty_col = type_col = del_col = val_col = -1
                start_row = -1
                
                # สแกนหาบรรทัดหัวตาราง
                for idx, row in df_bev.head(20).iterrows():
                    row_strs = [str(x).strip().lower() for x in row.values]
                    item_col = find_col_index(row_strs, item_names)
                    qty_col = find_col_index(row_strs, qty_names)
                    
                    if item_col != -1 and qty_col != -1:
                        start_row = idx
                        type_col = find_col_index(row_strs, type_names)
                        del_col = find_col_index(row_strs, del_names)
                        
                        amt_col = find_col_index(row_strs, amt_names)
                        net_col = find_col_index(row_strs, net_names)
                        val_col = amt_col if amt_col != -1 else net_col # ถ้าระบบมีคอลัมน์ยอดเงิน ให้ดึงยอดเงินก่อน
                        break
                        
                if start_row != -1:
                    for idx in range(start_row + 1, len(df_bev)):
                        row = df_bev.iloc[idx]
                        
                        # 1.1 ดึงจำนวนแก้ว
                        if pd.notna(row[item_col]):
                            item_name = str(row[item_col]).strip().lower()
                            if item_name and item_name != 'nan':
                                qty_str = str(row[qty_col]).replace(',', '') if qty_col != -1 and pd.notna(row[qty_col]) else "0"
                                try: qty = float(qty_str)
                                except: qty = 0
                                data_map[item_name] = data_map.get(item_name, 0) + qty

                        # 1.2 ดึงยอดขายแต่ละประเภท
                        if type_col != -1 and pd.notna(row[type_col]):
                            type_name = str(row[type_col]).strip().lower()
                            if type_name and type_name != 'nan':
                                val_str = str(row[val_col]).replace(',', '') if val_col != -1 and pd.notna(row[val_col]) else "0"
                                try: sale_val = float(val_str)
                                except: sale_val = 0
                                
                                if type_name in ['eatin', 'eat in', 'ทานที่ร้าน']:
                                    data_map['eatin'] = data_map.get('eatin', 0) + sale_val
                                    data_map['eat in'] = data_map.get('eat in', 0) + sale_val
                                elif type_name in ['takeaway', 'take away', 'สั่งกลับบ้าน']:
                                    data_map['takeaway'] = data_map.get('takeaway', 0) + sale_val
                                    data_map['take away'] = data_map.get('take away', 0) + sale_val
                                else:
                                    data_map[type_name] = data_map.get(type_name, 0) + sale_val
                                    
                                # ✨ จุดที่แก้บัค: ป้องกันบวกเบิ้ล ถ้าชื่อคอลัมน์ประเภทกับคอลัมน์แพลตฟอร์มดันเขียนเหมือนกัน
                                if del_col != -1 and pd.notna(row[del_col]):
                                    del_name = str(row[del_col]).strip().lower()
                                    if del_name and del_name != 'nan' and del_name != type_name:
                                        data_map[del_name] = data_map.get(del_name, 0) + sale_val
                else:
                    log_messages.append("⚠️ หาคอลัมน์ชื่อ 'Menu' และ 'Qty' ไม่เจอ")

                # --- 2. สกัดข้อมูลจากไฟล์ SaleReport (ดึงเฉพาะส่วนลด) ---
                df_sale = pd.read_csv(sale_file, header=None) if sale_file.name.endswith('.csv') else pd.read_excel(sale_file, header=None)
                discount_section = False
                
                for _, row in df_sale.iterrows():
                    col0 = str(row[0]).strip().lower()
                    if col0 == "discount summary" or col0 == "ส่วนลด":
                        discount_section = True
                        continue
                        
                    if discount_section:
                        if col0 in ["name", "ชื่อ", "", "nan", "none"]:
                            continue
                        amt_str = str(row[1]).replace(',', '') if pd.notna(row[1]) else "0"
                        try:
                            data_map[col0] = float(amt_str)
                        except:
                            pass

                # --- 3. นำข้อมูลไปหยอดลง Template ---
                wb = openpyxl.load_workbook(template_file)
                ws_bev = next((wb[sn] for sn in wb.sheetnames if 'bev' in sn.lower()), None)
                ws_sale = next((wb[sn] for sn in wb.sheetnames if 'sale' in sn.lower()), None)

                # 3.1 หยอดลงหน้า จำนวนแก้ว
                if ws_bev:
                    updated_bev = 0
                    for r in range(2, ws_bev.max_row + 1):
                        menu_cell = ws_bev.cell(row=r, column=2)
                        today_cell = ws_bev.cell(row=r, column=3)
                        yest_cell = ws_bev.cell(row=r, column=4)

                        if menu_cell.value and not str(menu_cell.value).startswith('='):
                            raw_name = str(menu_cell.value).strip()
                            m_name = raw_name.lower()
                            
                            if raw_name in ["Menu", "QTY", "Item", "Product"] or "Total" in raw_name:
                                continue
                            if today_cell.data_type == 'f' or yest_cell.data_type == 'f':
                                continue 

                            current_today = today_cell.value if isinstance(today_cell.value, (int, float)) else 0
                            yest_cell.value = current_today
                            
                            new_val = data_map.get(m_name, 0)
                            today_cell.value = new_val
                            if new_val > 0: updated_bev += 1
                                
                    log_messages.append(f"✅ อัปเดตยอด **จำนวนแก้ว** สำเร็จ ({updated_bev} เมนูที่มีคนสั่ง)")

                # 3.2 หยอดลงหน้า ยอดขายและส่วนลด
                if ws_sale:
                    updated_sale = 0
                    for r in range(2, ws_sale.max_row + 1):
                        cat_cell = ws_sale.cell(row=r, column=1)
                        today_cell = ws_sale.cell(row=r, column=2)
                        yest_cell = ws_sale.cell(row=r, column=4)

                        if cat_cell.value and not str(cat_cell.value).startswith('='):
                            raw_name = str(cat_cell.value).strip()
                            c_name = raw_name.lower()
                            
                            if raw_name in ["Category", "Sales - Discount", "Type"] or "Total" in raw_name or "Diff" in raw_name:
                                continue
                            if today_cell.data_type == 'f' or yest_cell.data_type == 'f':
                                continue 

                            current_today = today_cell.value if isinstance(today_cell.value, (int, float)) else 0
                            yest_cell.value = current_today
                            
                            new_val = data_map.get(c_name, 0)
                            today_cell.value = new_val
                            if new_val > 0 or current_today > 0: updated_sale += 1
                                
                    log_messages.append(f"✅ อัปเดต **ยอดขายและส่วนลด** สำเร็จ ({updated_sale} หมวดหมู่)")

                # --- 4. เตรียมไฟล์ Excel ให้ดาวน์โหลด ---
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.success("🎉 ประมวลผลเสร็จสมบูรณ์ ยอดตรงเป๊ะแล้วครับ!")
                for msg in log_messages:
                    st.write(msg)

                st.download_button(
                    label="📥 กดดาวน์โหลดไฟล์ Excel ตรงนี้ครับ",
                    data=output,
                    file_name="SamRoasters_Daily_Report_Updated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"❌ เกิดข้อผิดพลาด: {e}")
